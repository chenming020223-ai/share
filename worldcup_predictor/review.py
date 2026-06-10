from __future__ import annotations

import json
import re
from io import BytesIO
from statistics import mean
from typing import Any

from .betting import _asian_handicap_settlement, _asian_total_settlement
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .backtest import (
    _settle_handicap,
    _settle_match_winner,
    _settle_total,
    actual_result_key,
    brier_score,
    log_loss,
)
from .localization import localize_selection, translate_league_display, translate_team_display
from .paper_bankroll import build_paper_bankroll_timeline
from .poisson import poisson_pmf, score_matrix
from .storage import connect, official_batch_for_date


RESULT_LABELS = {"home_win": "主胜", "draw": "平局", "away_win": "客胜"}
MARKET_KEYS = ("home_win", "draw", "away_win")
SCORE_MATRIX_MAX_GOALS = 8


def build_daily_review(
    *,
    date: str,
    batch_id: int | None = None,
    use_official: bool = True,
    db_path: str | None = None,
) -> dict[str, Any]:
    batch_payload = _load_batch_payload(date=date, batch_id=batch_id, use_official=use_official, db_path=db_path)
    run_rows = _load_run_rows(date=date, batch_payload=batch_payload, db_path=db_path)
    results = _load_results([str(row["match_id"]) for row in run_rows], db_path=db_path)

    settled: list[dict[str, Any]] = []
    pending: list[dict[str, Any]] = []
    ev_candidates: list[dict[str, Any]] = []
    score_rows: list[dict[str, Any]] = []
    score_market_rows: list[dict[str, Any]] = []
    for row in run_rows:
        payload = json.loads(row["payload_json"])
        result = results.get(str(row["match_id"]))
        review_row = _review_row(row, payload, result)
        if review_row["settlementStatus"] == "已结算":
            settled.append(review_row)
            score_row = _score_distribution_row(row, payload, result)
            if score_row:
                score_rows.append(score_row)
            score_market_rows.extend(_score_market_attribution_rows(row, payload, result))
        else:
            pending.append(review_row)
        ev_candidates.extend(_ev_candidate_rows(row, payload, result))

    high_ev_anomalies = _high_ev_anomaly_rows(ev_candidates)
    market_line_backtest = _market_line_backtest(ev_candidates)
    score_distribution_backtest = _score_distribution_backtest(score_rows, score_market_rows)
    bankroll_timeline = build_paper_bankroll_timeline(db_path=db_path)
    summary = _review_summary(settled, pending, ev_candidates, batch_payload, bankroll_timeline, high_ev_anomalies)
    return {
        "date": date,
        "source": "official_batch" if batch_payload else "date_latest_runs",
        "batch": _batch_summary_payload(batch_payload),
        "summary": summary,
        "bankrollTimeline": bankroll_timeline,
        "settled": settled,
        "pending": pending,
        "evCandidates": ev_candidates,
        "highEvAnomalies": high_ev_anomalies,
        "evAnomalyGroups": _ev_anomaly_groups(high_ev_anomalies),
        "marketLineBacktest": market_line_backtest,
        "scoreDistributionBacktest": score_distribution_backtest,
        "gatePerformance": _group_performance(settled, "riskGate"),
        "qualityPerformance": _group_performance(settled, "qualityLabel"),
        "thresholdPerformance": _threshold_performance(settled),
        "notes": _review_notes(summary, batch_payload),
    }


def build_daily_review_excel(review: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "复盘摘要"
    for row in _summary_rows(review):
        ws.append(row)
    _write_sheet(
        wb,
        "已结算比赛",
        review.get("settled") or [],
        [
            "fixtureId",
            "runId",
            "match",
            "league",
            "apiLeague",
            "leagueCountry",
            "kickoffBeijing",
            "score90",
            "actualLabel",
            "topPredictionLabel",
            "topProbability",
            "hitLabel",
            "brier",
            "logLoss",
            "dataQuality",
            "qualityLabel",
            "riskGate",
            "simAction",
            "formalStake",
        ],
    )
    _write_sheet(
        wb,
        "待结算比赛",
        review.get("pending") or [],
        [
            "fixtureId",
            "runId",
            "match",
            "league",
            "kickoffBeijing",
            "topPredictionLabel",
            "topProbability",
            "dataQuality",
            "qualityLabel",
            "riskGate",
            "simAction",
        ],
    )
    _write_sheet(
        wb,
        "EV候选结算",
        review.get("evCandidates") or [],
        [
            "fixtureId",
            "runId",
            "match",
            "score90",
            "market",
            "selection",
            "selectionDisplay",
            "line",
            "odds",
            "modelProbability",
            "modelProbabilityLabel",
            "marketProbability",
            "expectedValue",
            "conservativeExpectedValue",
            "evPbaseResearch",
            "evPfinalExec",
            "evLayer",
            "signalStatus",
            "oddsBucket",
            "divergenceScore",
            "dataQuality",
            "league",
            "anomalyType",
            "actualNetPerUnit",
            "status",
            "action",
            "riskFlag",
            "reason",
        ],
    )
    _write_sheet(
        wb,
        "高EV异常",
        review.get("highEvAnomalies") or [],
        [
            "fixtureId",
            "runId",
            "match",
            "league",
            "score90",
            "market",
            "selection",
            "selectionDisplay",
            "line",
            "odds",
            "oddsBucket",
            "expectedValue",
            "actualNetPerUnit",
            "divergenceScore",
            "dataQuality",
            "anomalyType",
            "riskFlag",
        ],
    )
    _write_sheet(wb, "EV异常分组", review.get("evAnomalyGroups") or [])
    _write_sheet(wb, "盘口专项市场", (review.get("marketLineBacktest") or {}).get("marketGroups") or [])
    _write_sheet(wb, "盘口专项线别", (review.get("marketLineBacktest") or {}).get("lineGroups") or [])
    _write_sheet(wb, "比分分布比赛", (review.get("scoreDistributionBacktest") or {}).get("scoreRows") or [])
    _write_sheet(wb, "比分矩阵盘口", (review.get("scoreDistributionBacktest") or {}).get("marketRows") or [])
    _write_sheet(wb, "资金事件", (review.get("bankrollTimeline") or {}).get("events") or [])
    _write_sheet(wb, "闸门表现", review.get("gatePerformance") or [])
    _write_sheet(wb, "概率门槛", review.get("thresholdPerformance") or [])
    _write_sheet(wb, "质量分组", review.get("qualityPerformance") or [])
    _write_sheet(
        wb,
        "说明",
        [{"item": "口径", "value": note} for note in (review.get("notes") or [])],
        ["item", "value"],
    )
    _style_workbook(wb)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _load_batch_payload(
    *,
    date: str,
    batch_id: int | None,
    use_official: bool,
    db_path: str | None,
) -> dict[str, Any] | None:
    if batch_id:
        with connect(db_path) as conn:
            row = conn.execute("SELECT payload_json FROM batch_runs WHERE id = ?", (batch_id,)).fetchone()
        if not row:
            return None
        payload = json.loads(row["payload_json"])
        payload["batchRunId"] = batch_id
        return payload
    if use_official:
        return official_batch_for_date(date, db_path=db_path)
    return None


def _load_run_rows(
    *,
    date: str,
    batch_payload: dict[str, Any] | None,
    db_path: str | None,
) -> list[Any]:
    run_ids = _batch_run_ids(batch_payload)
    with connect(db_path) as conn:
        if run_ids:
            placeholders = ",".join("?" for _ in run_ids)
            return conn.execute(
                f"""
                SELECT *
                FROM prediction_runs
                WHERE id IN ({placeholders})
                ORDER BY id ASC
                """,
                run_ids,
            ).fetchall()
        rows = conn.execute(
            """
            SELECT *
            FROM prediction_runs
            WHERE date(datetime(created_at, '+8 hours')) = ?
            ORDER BY id ASC
            """,
            (date,),
        ).fetchall()
    latest: dict[str, Any] = {}
    for row in rows:
        latest[str(row["match_id"])] = row
    return list(latest.values())


def _load_results(fixture_ids: list[str], db_path: str | None) -> dict[str, Any]:
    if not fixture_ids:
        return {}
    placeholders = ",".join("?" for _ in fixture_ids)
    with connect(db_path) as conn:
        rows = conn.execute(
            f"SELECT * FROM match_results WHERE fixture_id IN ({placeholders})",
            fixture_ids,
        ).fetchall()
    return {str(row["fixture_id"]): row for row in rows}


def _review_row(row: Any, payload: dict[str, Any], result: Any | None) -> dict[str, Any]:
    match = payload.get("match") or {}
    meta = payload.get("meta") or {}
    probabilities = payload.get("probabilities") or {}
    display = probabilities.get("display") or probabilities.get("final") or {}
    pbase = probabilities.get("pbase") or probabilities.get("model") or {}
    qmkt = probabilities.get("qmkt") or probabilities.get("market") or {}
    top_key, top_probability = _top_probability(display)
    actual_key = None
    score = "-"
    brier = None
    loss = None
    pbase_brier = None
    pbase_loss = None
    qmkt_brier = None
    qmkt_loss = None
    hit = None
    if result:
        home_goals = int(result["home_goals_90"])
        away_goals = int(result["away_goals_90"])
        score = f"{home_goals}-{away_goals}"
        actual_key = actual_result_key(home_goals, away_goals)
        hit = top_key == actual_key
        if display:
            brier = brier_score(display, actual_key)
            loss = log_loss(display, actual_key)
        if pbase:
            pbase_brier = brier_score(pbase, actual_key)
            pbase_loss = log_loss(pbase, actual_key)
        if qmkt:
            qmkt_brier = brier_score(qmkt, actual_key)
            qmkt_loss = log_loss(qmkt, actual_key)
    quality = payload.get("dataQuality") or {}
    portfolio = payload.get("portfolio") or {}
    return {
        "fixtureId": str(row["match_id"]),
        "runId": int(row["id"]),
        "match": _match_label(match, row),
        "league": meta.get("leagueNameZh") or translate_league_display(meta.get("leagueName"), meta.get("leagueCountry")),
        "apiLeague": meta.get("leagueName") or "",
        "leagueCountry": meta.get("leagueCountry") or "",
        "kickoffBeijing": meta.get("kickoffBeijing") or "",
        "score90": score,
        "actual": actual_key,
        "actualLabel": RESULT_LABELS.get(actual_key, "待结算"),
        "topPrediction": top_key,
        "topPredictionLabel": RESULT_LABELS.get(top_key, "-"),
        "topProbability": top_probability,
        "hit": hit,
        "hitLabel": "命中" if hit else "未命中" if hit is False else "待结算",
        "settlementStatus": "已结算" if result else "待结算",
        "displayProbabilities": display,
        "pbaseProbabilities": pbase,
        "marketProbabilities": qmkt,
        "brier": brier,
        "logLoss": loss,
        "pbaseBrier": pbase_brier,
        "pbaseLogLoss": pbase_loss,
        "marketBrier": qmkt_brier,
        "marketLogLoss": qmkt_loss,
        "dataQuality": quality.get("score"),
        "qualityLabel": quality.get("gradeLabel") or quality.get("grade") or "-",
        "riskGate": _risk_gate(payload),
        "simAction": _sim_action(payload),
        "formalStake": float(portfolio.get("total_stake") or 0.0),
        "formalExpectedProfit": float(portfolio.get("expected_profit") or 0.0),
    }


def _ev_candidate_rows(row: Any, payload: dict[str, Any], result: Any | None) -> list[dict[str, Any]]:
    match = payload.get("match") or {}
    meta = payload.get("meta") or {}
    quality = payload.get("dataQuality") or {}
    score = "-"
    home_goals = away_goals = None
    if result:
        home_goals = int(result["home_goals_90"])
        away_goals = int(result["away_goals_90"])
        score = f"{home_goals}-{away_goals}"
    rows: list[dict[str, Any]] = []
    for recommendation in payload.get("recommendations") or []:
        if not isinstance(recommendation, dict):
            continue
        ev = _number(
            recommendation.get("ev_pbase_research")
            if recommendation.get("ev_pbase_research") is not None
            else (
                recommendation.get("expected_value_per_unit")
                if recommendation.get("expected_value_per_unit") is not None
                else recommendation.get("audit_expected_value_per_unit")
            )
        )
        if ev is None:
            continue
        conservative_ev = _number(
            recommendation.get("conservative_ev_pbase_research")
            if recommendation.get("conservative_ev_pbase_research") is not None
            else (
                recommendation.get("conservative_expected_value_per_unit")
                if recommendation.get("conservative_expected_value_per_unit") is not None
                else recommendation.get("audit_conservative_expected_value_per_unit")
            )
        )
        signal_status = str(recommendation.get("signal_status") or "")
        ev_layer = str(recommendation.get("ev_layer") or "")
        net = None
        if home_goals is not None and away_goals is not None:
            net = _settle_research_candidate(payload, recommendation, home_goals, away_goals)
        market = str(recommendation.get("market") or "")
        odds = _number(recommendation.get("odds"))
        model_probability = _number(recommendation.get("model_probability"))
        model_probability_label = str(
            recommendation.get("model_probability_label")
            or ("模型胜率" if market == "胜平负" else "正收益概率")
        )
        market_probability = _number(recommendation.get("market_probability"))
        divergence = (
            abs(model_probability - market_probability)
            if model_probability is not None and market_probability is not None
            else None
        )
        flag = _ev_risk_flag(market, odds, ev, net, recommendation)
        anomaly_type = _ev_anomaly_type(market, odds, ev, net, divergence)
        rows.append(
            {
                "fixtureId": str(row["match_id"]),
                "runId": int(row["id"]),
                "match": _match_label(match, row),
                "league": meta.get("leagueNameZh") or translate_league_display(meta.get("leagueName"), meta.get("leagueCountry")),
                "score90": score,
                "market": market,
                "selection": str(recommendation.get("selection") or ""),
                "selectionDisplay": _selection_display(row, match, market, recommendation),
                "line": recommendation.get("line"),
                "odds": odds,
                "oddsBucket": _odds_bucket(odds),
                "modelProbability": model_probability,
                "modelProbabilityLabel": model_probability_label,
                "marketProbability": market_probability,
                "divergenceScore": divergence,
                "dataQuality": _number(quality.get("score")),
                "expectedValue": ev,
                "conservativeExpectedValue": conservative_ev,
                "evPbaseResearch": ev,
                "evPfinalExec": _number(recommendation.get("ev_pfinal_exec")),
                "evLayer": ev_layer,
                "signalStatus": signal_status,
                "anomalyType": anomaly_type,
                "actualNetPerUnit": net,
                "status": str(recommendation.get("ev_status") or ""),
                "action": str(recommendation.get("action") or ""),
                "riskFlag": flag,
                "reason": str(recommendation.get("reason") or ""),
            }
        )
    return rows


def _review_summary(
    settled: list[dict[str, Any]],
    pending: list[dict[str, Any]],
    ev_candidates: list[dict[str, Any]],
    batch_payload: dict[str, Any] | None,
    bankroll_timeline: dict[str, Any] | None = None,
    high_ev_anomalies: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    hit_count = sum(1 for row in settled if row["hit"])
    formal_stake = sum(row.get("formalStake") or 0.0 for row in settled + pending)
    ev_settled = [row for row in ev_candidates if row.get("actualNetPerUnit") is not None]
    ev_loss_count = sum(1 for row in ev_settled if (row.get("actualNetPerUnit") or 0.0) < 0)
    high_ev_losses = [
        row for row in ev_settled
        if (row.get("expectedValue") or 0.0) >= 0.20 and (row.get("actualNetPerUnit") or 0.0) < 0
    ]
    bankroll_summary = (bankroll_timeline or {}).get("summary") or {}
    return {
        "totalMatches": len(settled) + len(pending),
        "settledMatches": len(settled),
        "pendingMatches": len(pending),
        "hitCount": hit_count,
        "hitRate": hit_count / len(settled) if settled else None,
        "avgBrier": _avg(row.get("brier") for row in settled),
        "avgLogLoss": _avg(row.get("logLoss") for row in settled),
        "avgPbaseBrier": _avg(row.get("pbaseBrier") for row in settled),
        "avgPbaseLogLoss": _avg(row.get("pbaseLogLoss") for row in settled),
        "avgMarketBrier": _avg(row.get("marketBrier") for row in settled),
        "avgMarketLogLoss": _avg(row.get("marketLogLoss") for row in settled),
        "formalStake": formal_stake,
        "formalSignalState": "正式EV关闭，资金占用为零" if formal_stake <= 0 else "存在纸上模拟资金占用",
        "evCandidateCount": len(ev_candidates),
        "settledEvCandidateCount": len(ev_settled),
        "settledEvNetPerUnit": sum(row.get("actualNetPerUnit") or 0.0 for row in ev_settled),
        "settledEvLossCount": ev_loss_count,
        "highEvLossCount": len(high_ev_losses),
        "highEvAnomalyCount": len(high_ev_anomalies or []),
        "paperCash": bankroll_summary.get("cash"),
        "paperReservedStake": bankroll_summary.get("reservedStake"),
        "paperEquity": bankroll_summary.get("equity"),
        "paperDrawdownPct": bankroll_summary.get("drawdownPct"),
        "paperRiskLabel": bankroll_summary.get("riskLabel"),
        "batchRunId": (batch_payload or {}).get("batchRunId"),
        "isOfficialBatch": bool((batch_payload or {}).get("isOfficial")),
    }


def _threshold_performance(settled: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for threshold in (0.50, 0.55, 0.60, 0.65, 0.70):
        items = [row for row in settled if (row.get("topProbability") or 0.0) >= threshold]
        hits = sum(1 for row in items if row.get("hit"))
        rows.append(
            {
                "threshold": threshold,
                "matches": len(items),
                "hits": hits,
                "hitRate": hits / len(items) if items else None,
            }
        )
    return rows


def _group_performance(settled: list[dict[str, Any]], key: str) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in settled:
        groups.setdefault(str(row.get(key) or "-"), []).append(row)
    output: list[dict[str, Any]] = []
    for label, items in sorted(groups.items()):
        hits = sum(1 for row in items if row.get("hit"))
        output.append(
            {
                "group": label,
                "matches": len(items),
                "hits": hits,
                "hitRate": hits / len(items) if items else None,
                "avgBrier": _avg(row.get("brier") for row in items),
                "avgLogLoss": _avg(row.get("logLoss") for row in items),
            }
        )
    return output


def _review_notes(summary: dict[str, Any], batch_payload: dict[str, Any] | None) -> list[str]:
    notes = [
        "复盘口径：每个 fixture 只取本次复盘范围内的一条预测；已完赛使用 match_results 中的 90 分钟比分。",
        "研究 EV 候选用于诊断模型与盘口分歧，不等同于正式下注信号。",
        "正式 EV 仍以 pfinal 通过校准验收为前提；当前复盘不会自动开放 BUY。",
    ]
    if batch_payload:
        label = "官方批次" if batch_payload.get("isOfficial") else "指定批次"
        notes.insert(0, f"复盘来源：{label} #{batch_payload.get('batchRunId')}。")
    if summary.get("highEvLossCount"):
        notes.append(f"发现 {summary['highEvLossCount']} 个高 EV 但实际亏损的候选，需优先复核冷门和大小球概率。")
    return notes


def _settle_research_candidate(
    payload: dict[str, Any],
    recommendation: dict[str, Any],
    home_goals: int,
    away_goals: int,
) -> float | None:
    market = str(recommendation.get("market") or "")
    selection = str(recommendation.get("selection") or "")
    odds = _number(recommendation.get("odds"))
    line = recommendation.get("line")
    if odds is None or odds <= 1:
        return None
    try:
        if market == "胜平负":
            return _settle_match_winner(payload, selection, home_goals, away_goals, odds)
        if market == "大小球":
            return _settle_total(selection, float(line), home_goals, away_goals, odds)
        if market == "让球":
            return _settle_handicap(payload, selection, float(line), home_goals, away_goals, odds)
    except (TypeError, ValueError):
        return None
    return None


def _ev_risk_flag(
    market: str,
    odds: float | None,
    ev: float,
    net: float | None,
    recommendation: dict[str, Any],
) -> str:
    flags: list[str] = []
    signal_status = str(recommendation.get("signal_status") or "")
    ev_layer = str(recommendation.get("ev_layer") or "")
    if signal_status == "SUSPENDED":
        flags.append("模拟暂停")
    if signal_status == "RESEARCH_WATCH":
        flags.append("研究观察")
    if ev_layer and ev_layer != "pfinal_exec":
        flags.append("研究层EV")
    if recommendation.get("ev_status") == "SUSPENDED_MODEL_DIVERGENCE":
        flags.append("模型分歧暂停")
    if ev >= 0.20:
        flags.append("高EV复核")
    if odds is not None and odds >= 5.0:
        flags.append("高赔冷门")
    if market == "大小球" and ev >= 0.15:
        flags.append("大小球高EV")
    if net is not None and ev >= 0.20 and net < 0:
        flags.append("高EV实际亏损")
    return "、".join(flags) or "常规复核"


def _ev_anomaly_type(
    market: str,
    odds: float | None,
    ev: float,
    net: float | None,
    divergence: float | None,
) -> str:
    if net is not None and ev >= 0.20 and net < 0:
        return "高EV实际亏损"
    if odds is not None and odds >= 5.0:
        return "高赔冷门复核"
    if market == "大小球" and ev >= 0.15:
        return "大小球高EV复核"
    if divergence is not None and divergence >= 0.12:
        return "模型市场分歧复核"
    if ev >= 0.15:
        return "高EV复核"
    return "常规复核"


def _odds_bucket(odds: float | None) -> str:
    if odds is None:
        return "-"
    if odds < 1.8:
        return "<1.80"
    if odds < 2.5:
        return "1.80-2.49"
    if odds < 4.0:
        return "2.50-3.99"
    if odds < 6.0:
        return "4.00-5.99"
    return ">=6.00"


def _high_ev_anomaly_rows(ev_candidates: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in ev_candidates:
        ev = row.get("expectedValue") or 0.0
        odds = row.get("odds") or 0.0
        net = row.get("actualNetPerUnit")
        divergence = row.get("divergenceScore") or 0.0
        include = (
            net is not None
            and (
                ev >= 0.15
                or odds >= 5.0
                or divergence >= 0.12
                or (net < 0 and ev >= 0.08)
            )
        )
        if not include:
            continue
        risk_score = ev * 2.0 + divergence + (0.4 if odds >= 5.0 else 0.0) + (0.6 if net is not None and net < 0 else 0.0)
        rows.append({**row, "riskScore": risk_score})
    return sorted(rows, key=lambda item: item.get("riskScore") or 0.0, reverse=True)


def _ev_anomaly_groups(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        groups.setdefault((str(row.get("market") or "-"), str(row.get("oddsBucket") or "-")), []).append(row)
    output: list[dict[str, Any]] = []
    for (market, odds_bucket), items in groups.items():
        losses = [item for item in items if (item.get("actualNetPerUnit") or 0.0) < 0]
        output.append(
            {
                "market": market,
                "oddsBucket": odds_bucket,
                "count": len(items),
                "lossCount": len(losses),
                "lossRate": len(losses) / len(items) if items else None,
                "avgEv": _avg(item.get("expectedValue") for item in items),
                "avgDivergence": _avg(item.get("divergenceScore") for item in items),
                "netPerUnit": sum(item.get("actualNetPerUnit") or 0.0 for item in items),
            }
        )
    return sorted(output, key=lambda item: (item["lossCount"], item["count"], abs(item["netPerUnit"])), reverse=True)


def _market_line_backtest(ev_candidates: list[dict[str, Any]]) -> dict[str, Any]:
    rows = [
        {**row, "lineKey": _market_line_key(row), "selectionSide": _selection_side(row)}
        for row in ev_candidates
        if row.get("market") in {"大小球", "让球"}
        and row.get("actualNetPerUnit") is not None
        and _market_line_key(row) != "-"
    ]
    return {
        "summary": {
            "settledCandidates": len(rows),
            "marketCount": len({row.get("market") for row in rows}),
            "lineCount": len({(row.get("market"), row.get("lineKey")) for row in rows}),
            "totalNetPerUnit": sum(row.get("actualNetPerUnit") or 0.0 for row in rows),
            "highEvLossCount": sum(
                1 for row in rows
                if (row.get("expectedValue") or 0.0) >= 0.15 and (row.get("actualNetPerUnit") or 0.0) < 0
            ),
            "approvalStatus": "research_only",
            "approvalLabel": "仅供研究，不能开放执行",
        },
        "marketGroups": _aggregate_market_line_rows(rows, ("market",)),
        "lineGroups": _aggregate_market_line_rows(rows, ("market", "lineKey")),
        "selectionGroups": _aggregate_market_line_rows(rows, ("market", "lineKey", "selectionSide")),
        "notes": [
            "大小球与让球专项回测基于已结算研究 EV 候选，仍使用 90 分钟比分结算。",
            "亚洲盘的 modelProbability 是正收益概率，不等同于简单胜率；正式准入必须看等额盈亏、线别稳定性和样本量。",
            "样本不足时只输出诊断，不得据此开放 PAPER_BUY。",
        ],
    }


def _aggregate_market_line_rows(rows: list[dict[str, Any]], keys: tuple[str, ...]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, ...], list[dict[str, Any]]] = {}
    for row in rows:
        groups.setdefault(tuple(str(row.get(key) or "-") for key in keys), []).append(row)
    output: list[dict[str, Any]] = []
    for key_values, items in groups.items():
        count = len(items)
        positive = [item for item in items if (item.get("actualNetPerUnit") or 0.0) > 0]
        push = [item for item in items if abs(item.get("actualNetPerUnit") or 0.0) <= 1e-9]
        losses = [item for item in items if (item.get("actualNetPerUnit") or 0.0) < 0]
        net = sum(item.get("actualNetPerUnit") or 0.0 for item in items)
        high_ev_losses = [
            item for item in items
            if (item.get("expectedValue") or 0.0) >= 0.15 and (item.get("actualNetPerUnit") or 0.0) < 0
        ]
        market = key_values[0] if len(key_values) >= 1 else "-"
        line_key = key_values[1] if len(key_values) >= 2 else "全部"
        side = key_values[2] if len(key_values) >= 3 else "全部"
        status, label, diagnosis = _line_backtest_status(count, net, len(losses), len(high_ev_losses))
        output.append(
            {
                "market": market,
                "lineKey": line_key,
                "selectionSide": side,
                "count": count,
                "positiveCount": len(positive),
                "pushCount": len(push),
                "lossCount": len(losses),
                "positiveRate": len(positive) / count if count else None,
                "lossRate": len(losses) / count if count else None,
                "avgModelProbability": _avg(item.get("modelProbability") for item in items),
                "avgMarketProbability": _avg(item.get("marketProbability") for item in items),
                "avgEv": _avg(item.get("expectedValue") for item in items),
                "avgConservativeEv": _avg(item.get("conservativeExpectedValue") for item in items),
                "avgDivergence": _avg(item.get("divergenceScore") for item in items),
                "avgDataQuality": _avg(item.get("dataQuality") for item in items),
                "netPerUnit": net,
                "roiPerUnit": net / count if count else None,
                "highEvLossCount": len(high_ev_losses),
                "status": status,
                "statusLabel": label,
                "diagnosis": diagnosis,
            }
        )
    return sorted(
        output,
        key=lambda item: (
            item["status"] != "UNSTABLE",
            item["status"] != "HIGH_EV_RISK",
            -item["count"],
            item["netPerUnit"],
        ),
    )


def _line_backtest_status(count: int, net: float, loss_count: int, high_ev_loss_count: int) -> tuple[str, str, str]:
    if count < 5:
        return "SAMPLE_TOO_SMALL", "样本不足", "已结算样本少于 5 条，仅供观察，不能形成准入结论。"
    loss_rate = loss_count / count if count else 0.0
    if net < 0 and loss_rate >= 0.50:
        return "UNSTABLE", "亏损集中", "该市场/盘口线等额回测为负且亏损占比较高，应优先降权或暂停。"
    if high_ev_loss_count > 0:
        return "HIGH_EV_RISK", "高EV风险", "存在高 EV 候选实际亏损，需复核比分分布和盘口线选择。"
    if net > 0:
        return "OBSERVE_POSITIVE", "观察为正", "当前等额结果为正，但仍需时间切分和更多样本验证。"
    return "RESEARCH_WATCH", "继续观察", "暂无明显正向或负向证据，保持研究观察。"


def _line_key(value: Any) -> str:
    number = _number(value)
    plus_prefix = False
    if number is None and isinstance(value, str):
        matches = re.findall(r"[+-]?\d+(?:\.\d+)?", value)
        if "/" in value and len(matches) >= 2:
            numbers = [_number(item) for item in matches[-2:]]
            if all(item is not None for item in numbers):
                number = sum(numbers) / 2
                plus_prefix = any(item.strip().startswith("+") for item in matches[-2:]) and number > 0
        elif matches:
            number = _number(matches[-1])
            plus_prefix = matches[-1].strip().startswith("+") and (number or 0) > 0
    if number is None:
        return "-"
    if plus_prefix:
        return f"+{number:g}"
    return f"{number:g}"


def _market_line_key(row: dict[str, Any]) -> str:
    market = str(row.get("market") or "")
    if market in {"大小球", "让球"}:
        selection_key = _line_key(row.get("selection"))
        if selection_key != "-":
            return selection_key
    return _line_key(row.get("line"))


def _selection_display(row: Any, match: dict[str, Any], market: str, recommendation: dict[str, Any]) -> str:
    selection = str(recommendation.get("selection") or "")
    home_raw = str(match.get("home") or row["home_team"] or "")
    away_raw = str(match.get("away") or row["away_team"] or "")
    home_label = str(match.get("homeZh") or translate_team_display(home_raw, "主队"))
    away_label = str(match.get("awayZh") or translate_team_display(away_raw, "客队"))
    localized = localize_selection(selection, home_raw, away_raw)
    line_key = _market_line_key({"market": market, "selection": selection, "line": recommendation.get("line")})
    if market == "大小球" and line_key != "-":
        text = selection.strip().casefold()
        if selection.startswith("大") or text.startswith("over"):
            return f"大 {line_key}"
        if selection.startswith("小") or text.startswith("under"):
            return f"小 {line_key}"
        return localized or f"大小球 {line_key}"
    if market == "让球" and line_key != "-":
        selection_fold = selection.casefold()
        home_candidates = {home_raw.casefold(), home_label.casefold()}
        away_candidates = {away_raw.casefold(), away_label.casefold()}
        if any(candidate and candidate in selection_fold for candidate in home_candidates):
            return f"{home_label} {line_key}"
        if any(candidate and candidate in selection_fold for candidate in away_candidates):
            return f"{away_label} {line_key}"
    return localized or selection or "-"


def _selection_side(row: dict[str, Any]) -> str:
    market = str(row.get("market") or "")
    selection = str(row.get("selection") or "")
    if market == "大小球":
        if selection.startswith("大"):
            return "大"
        if selection.startswith("小"):
            return "小"
    if market == "让球":
        if "+" in selection:
            return "加号方向"
        if "-" in selection:
            return "减号方向"
        return "让球方向"
    return "-"


def _score_distribution_backtest(score_rows: list[dict[str, Any]], market_rows: list[dict[str, Any]]) -> dict[str, Any]:
    settled = len(score_rows)
    top1_hits = sum(1 for row in score_rows if row.get("actualScoreRank") == 1)
    top6_hits = sum(1 for row in score_rows if (row.get("actualScoreRank") or 999) <= 6)
    underestimates = sum(1 for row in score_rows if (row.get("totalGoalError") or 0.0) >= 0.75)
    overestimates = sum(1 for row in score_rows if (row.get("totalGoalError") or 0.0) <= -0.75)
    high_ev_losses = [
        row for row in market_rows
        if (row.get("expectedValueFromMatrix") or 0.0) >= 0.15 and (row.get("actualNetPerUnit") or 0.0) < 0
    ]
    return {
        "summary": {
            "settledMatches": settled,
            "marketAttributionCount": len(market_rows),
            "avgActualScoreProbability": _avg(row.get("actualScoreProbability") for row in score_rows),
            "top1HitRate": top1_hits / settled if settled else None,
            "top6HitRate": top6_hits / settled if settled else None,
            "avgExpectedTotalGoals": _avg(row.get("expectedTotalGoals") for row in score_rows),
            "avgActualTotalGoals": _avg(row.get("actualTotalGoals") for row in score_rows),
            "avgTotalGoalError": _avg(row.get("totalGoalError") for row in score_rows),
            "avgAbsTotalGoalError": _avg(row.get("absTotalGoalError") for row in score_rows),
            "underestimateCount": underestimates,
            "overestimateCount": overestimates,
            "avgTailMass": _avg(row.get("tailMass") for row in score_rows),
            "highEvLossCount": len(high_ev_losses),
            "approvalStatus": "research_only",
            "approvalLabel": "比分分布研究审计，不能开放执行",
        },
        "statusGroups": _aggregate_score_distribution_rows(score_rows),
        "scoreRows": sorted(score_rows, key=lambda row: (row.get("statusPriority") or 0, row.get("absTotalGoalError") or 0.0), reverse=True),
        "marketRows": sorted(
            market_rows,
            key=lambda row: (
                row.get("statusPriority") or 0,
                abs(row.get("evRebuildDelta") or 0.0),
                row.get("expectedValueFromMatrix") or 0.0,
            ),
            reverse=True,
        ),
        "notes": [
            "比分分布审计使用当时保存的预期进球重建 pbase 泊松比分矩阵。",
            "该模块用于定位大小球 / 让球 EV 失真来源，不代表 pfinal 已通过验收。",
            "若总进球长期偏高或偏低，下一步应校准预期进球和比分矩阵尾部，而不是直接放开模拟舱。",
        ],
    }


def _score_distribution_row(row: Any, payload: dict[str, Any], result: Any | None) -> dict[str, Any] | None:
    if not result:
        return None
    expected = payload.get("expectedGoals") or {}
    home_mu = _number(expected.get("home"))
    away_mu = _number(expected.get("away"))
    if home_mu is None or away_mu is None:
        return None
    home_goals = int(result["home_goals_90"])
    away_goals = int(result["away_goals_90"])
    matrix = score_matrix(home_mu, away_mu, SCORE_MATRIX_MAX_GOALS)
    actual_score = f"{home_goals}-{away_goals}"
    actual_probability = matrix.get((home_goals, away_goals), 0.0)
    rank = _score_rank(matrix, home_goals, away_goals)
    actual_total = home_goals + away_goals
    expected_total = home_mu + away_mu
    total_error = actual_total - expected_total
    actual_key = actual_result_key(home_goals, away_goals)
    probabilities = payload.get("probabilities") or {}
    pbase = probabilities.get("pbase") or probabilities.get("model") or {}
    status, label, priority = _score_distribution_status(actual_probability, rank, total_error, home_goals - home_mu, away_goals - away_mu)
    meta = payload.get("meta") or {}
    match = payload.get("match") or {}
    return {
        "fixtureId": str(row["match_id"]),
        "runId": int(row["id"]),
        "match": _match_label(match, row),
        "league": meta.get("leagueNameZh") or translate_league_display(meta.get("leagueName"), meta.get("leagueCountry")),
        "score90": actual_score,
        "expectedGoalsHome": home_mu,
        "expectedGoalsAway": away_mu,
        "expectedTotalGoals": expected_total,
        "actualTotalGoals": actual_total,
        "totalGoalError": total_error,
        "absTotalGoalError": abs(total_error),
        "homeGoalError": home_goals - home_mu,
        "awayGoalError": away_goals - away_mu,
        "actualScoreProbability": actual_probability,
        "actualScoreRank": rank,
        "actualResultProbability": _number(pbase.get(actual_key)),
        "topScores": " / ".join(f"{score} {probability * 100:.1f}%" for score, probability in _top_scores_from_matrix(matrix, 3)),
        "lowTotalMass": _score_total_mass(matrix, max_total=1),
        "highTotalMass": _score_total_mass(matrix, min_total=4),
        "tailMass": _score_tail_mass(home_mu, away_mu, SCORE_MATRIX_MAX_GOALS),
        "status": status,
        "statusLabel": label,
        "statusPriority": priority,
    }


def _score_market_attribution_rows(row: Any, payload: dict[str, Any], result: Any | None) -> list[dict[str, Any]]:
    if not result:
        return []
    expected = payload.get("expectedGoals") or {}
    home_mu = _number(expected.get("home"))
    away_mu = _number(expected.get("away"))
    if home_mu is None or away_mu is None:
        return []
    home_goals = int(result["home_goals_90"])
    away_goals = int(result["away_goals_90"])
    matrix = score_matrix(home_mu, away_mu, SCORE_MATRIX_MAX_GOALS)
    rows: list[dict[str, Any]] = []
    for recommendation in payload.get("recommendations") or []:
        if not isinstance(recommendation, dict):
            continue
        market = str(recommendation.get("market") or "")
        if market not in {"大小球", "让球"}:
            continue
        odds = _number(recommendation.get("odds"))
        if odds is None or odds <= 1:
            continue
        line = _number(recommendation.get("line"))
        selection = str(recommendation.get("selection") or "")
        if market == "大小球":
            line = line if line is not None else _number(_line_key(selection))
            side = _total_side(selection)
            if line is None or side is None:
                continue
            settlement = _asian_total_settlement(matrix, line, side, odds)
        else:
            side = _handicap_side(payload, selection)
            if line is None or side is None:
                continue
            settlement = _asian_handicap_settlement(matrix, line, side, odds)
        actual_net = _settle_research_candidate(payload, recommendation, home_goals, away_goals)
        expected_ev = _number(settlement.get("ev"))
        rec_ev = _number(
            recommendation.get("ev_pbase_research")
            if recommendation.get("ev_pbase_research") is not None
            else (
                recommendation.get("expected_value_per_unit")
                if recommendation.get("expected_value_per_unit") is not None
                else recommendation.get("audit_expected_value_per_unit")
            )
        )
        status, label, priority = _score_market_status(expected_ev, actual_net, settlement, rec_ev)
        match = payload.get("match") or {}
        meta = payload.get("meta") or {}
        rows.append(
            {
                "fixtureId": str(row["match_id"]),
                "runId": int(row["id"]),
                "match": _match_label(match, row),
                "league": meta.get("leagueNameZh") or translate_league_display(meta.get("leagueName"), meta.get("leagueCountry")),
                "score90": f"{home_goals}-{away_goals}",
                "market": market,
                "selection": selection,
                "selectionDisplay": _selection_display(row, match, market, recommendation),
                "lineKey": _market_line_key({"market": market, "selection": selection, "line": line}),
                "odds": odds,
                "modelProbability": _number(recommendation.get("model_probability")),
                "matrixPositiveProbability": _number(settlement.get("positive")),
                "winFraction": _number(settlement.get("win_fraction")),
                "lossFraction": _number(settlement.get("loss_fraction")),
                "fullWinMass": _number(settlement.get("full_win")),
                "halfWinMass": _number(settlement.get("half_win")),
                "pushMass": _number(settlement.get("push")),
                "halfLossMass": _number(settlement.get("half_loss")),
                "fullLossMass": _number(settlement.get("full_loss")),
                "breakEvenOdds": _number(settlement.get("break_even_odds")),
                "expectedValueFromMatrix": expected_ev,
                "recommendationExpectedValue": rec_ev,
                "evRebuildDelta": (expected_ev - rec_ev) if expected_ev is not None and rec_ev is not None else None,
                "actualNetPerUnit": actual_net,
                "status": status,
                "statusLabel": label,
                "statusPriority": priority,
            }
        )
    return rows


def _aggregate_score_distribution_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        groups.setdefault(str(row.get("statusLabel") or "-"), []).append(row)
    output: list[dict[str, Any]] = []
    for label, items in groups.items():
        output.append(
            {
                "statusLabel": label,
                "count": len(items),
                "avgActualScoreProbability": _avg(item.get("actualScoreProbability") for item in items),
                "avgTotalGoalError": _avg(item.get("totalGoalError") for item in items),
                "avgAbsTotalGoalError": _avg(item.get("absTotalGoalError") for item in items),
                "avgTailMass": _avg(item.get("tailMass") for item in items),
            }
        )
    return sorted(output, key=lambda item: (item["count"], item["avgAbsTotalGoalError"] or 0.0), reverse=True)


def _score_distribution_status(
    actual_probability: float,
    rank: int | None,
    total_error: float,
    home_error: float,
    away_error: float,
) -> tuple[str, str, int]:
    if abs(total_error) >= 2.0:
        return "TOTAL_GOAL_MISS", "总进球严重偏差", 4
    if abs(home_error) >= 2.0 or abs(away_error) >= 2.0:
        return "TEAM_GOAL_MISS", "单队进球偏差", 3
    if actual_probability < 0.015 or (rank is not None and rank > 20):
        return "LOW_PROB_SCORE", "低概率比分", 2
    if abs(total_error) >= 1.0:
        return "TOTAL_GOAL_WATCH", "总进球偏差观察", 1
    return "NORMAL_RESEARCH", "正常观察", 0


def _score_market_status(
    expected_ev: float | None,
    actual_net: float | None,
    settlement: dict[str, float],
    rec_ev: float | None,
) -> tuple[str, str, int]:
    if expected_ev is not None and rec_ev is not None and abs(expected_ev - rec_ev) >= 0.03:
        return "EV_REBUILD_MISMATCH", "EV复算差异", 5
    if expected_ev is not None and expected_ev >= 0.15 and actual_net is not None and actual_net < 0:
        return "HIGH_EV_LOSS", "高EV实际亏损", 4
    loss_fraction = _number(settlement.get("loss_fraction")) or 0.0
    win_fraction = _number(settlement.get("win_fraction")) or 0.0
    if actual_net is not None and actual_net < 0 and loss_fraction >= win_fraction:
        return "LOSS_WEIGHT_HEAVY", "亏损权重偏高", 3
    if actual_net is not None and actual_net < 0:
        return "SINGLE_LOSS", "单场亏损", 2
    if actual_net is not None and actual_net > 0:
        return "SINGLE_WIN", "单场盈利", 1
    return "PUSH_OR_WATCH", "走水/观察", 0


def _total_side(selection: str) -> str | None:
    text = selection.strip().casefold()
    if selection.startswith("大") or text.startswith("over"):
        return "over"
    if selection.startswith("小") or text.startswith("under"):
        return "under"
    return None


def _handicap_side(payload: dict[str, Any], selection: str) -> str | None:
    match = payload.get("match") or {}
    selection_fold = selection.casefold()
    home_names = {
        str(match.get("home") or "").casefold(),
        str(match.get("homeZh") or "").casefold(),
    }
    away_names = {
        str(match.get("away") or "").casefold(),
        str(match.get("awayZh") or "").casefold(),
    }
    if any(name and selection_fold.startswith(name) for name in home_names):
        return "home"
    if any(name and selection_fold.startswith(name) for name in away_names):
        return "away"
    return None


def _score_rank(matrix: dict[tuple[int, int], float], home_goals: int, away_goals: int) -> int | None:
    ranked = sorted(matrix.items(), key=lambda item: item[1], reverse=True)
    for index, (score, _) in enumerate(ranked, start=1):
        if score == (home_goals, away_goals):
            return index
    return None


def _top_scores_from_matrix(matrix: dict[tuple[int, int], float], limit: int) -> list[tuple[str, float]]:
    ranked = sorted(matrix.items(), key=lambda item: item[1], reverse=True)
    return [(f"{home}-{away}", probability) for (home, away), probability in ranked[:limit]]


def _score_total_mass(
    matrix: dict[tuple[int, int], float],
    *,
    min_total: int | None = None,
    max_total: int | None = None,
) -> float:
    total = 0.0
    for (home_goals, away_goals), probability in matrix.items():
        goals = home_goals + away_goals
        if min_total is not None and goals < min_total:
            continue
        if max_total is not None and goals > max_total:
            continue
        total += probability
    return total


def _score_tail_mass(home_mu: float, away_mu: float, max_goals: int) -> float:
    retained = 0.0
    for home_goals in range(max_goals + 1):
        home_prob = poisson_pmf(home_goals, home_mu)
        for away_goals in range(max_goals + 1):
            retained += home_prob * poisson_pmf(away_goals, away_mu)
    return max(0.0, min(1.0, 1.0 - retained))


def _risk_gate(payload: dict[str, Any]) -> str:
    recommendations = [item for item in payload.get("recommendations") or [] if isinstance(item, dict)]
    statuses = {item.get("ev_status") for item in recommendations}
    signal_statuses = {item.get("signal_status") for item in recommendations}
    actions = {item.get("action") for item in recommendations}
    if "SUSPENDED" in signal_statuses:
        return "模拟暂停"
    if "SUSPENDED_MODEL_DIVERGENCE" in statuses:
        return "模型分歧暂停"
    if actions == {"NO_MARKET"} or (actions and actions.issubset({"NO_MARKET"})):
        return "市场缺失"
    return "研究观察"


def _sim_action(payload: dict[str, Any]) -> str:
    recommendations = [item for item in payload.get("recommendations") or [] if isinstance(item, dict)]
    signal_statuses = {str(item.get("signal_status") or "") for item in recommendations}
    active = [item for item in recommendations if item.get("signal_status") == "PAPER_BUY"]
    if active:
        return "纸上观察"
    if "SUSPENDED" in signal_statuses:
        return "暂停"
    if "MODEL_CANDIDATE" in signal_statuses:
        return "模型候选"
    if "RESEARCH_WATCH" in signal_statuses:
        return "研究观察"
    active = [item for item in recommendations if item.get("action") in {"BUY", "PAPER_BUY"}]
    if active:
        return "模型候选"
    if any(item.get("action") == "WATCH" for item in recommendations):
        return "观望"
    if any(item.get("action") == "NO_MARKET" for item in recommendations):
        return "市场缺失"
    return "-"


def _batch_run_ids(batch_payload: dict[str, Any] | None) -> list[int]:
    if not batch_payload:
        return []
    ids: list[int] = []
    for item in batch_payload.get("collected") or []:
        run_id = item.get("runId") if isinstance(item, dict) else None
        try:
            if run_id is not None:
                ids.append(int(run_id))
        except (TypeError, ValueError):
            continue
    return ids


def _batch_summary_payload(batch_payload: dict[str, Any] | None) -> dict[str, Any] | None:
    if not batch_payload:
        return None
    return {
        "batchRunId": batch_payload.get("batchRunId"),
        "date": batch_payload.get("date"),
        "scope": batch_payload.get("scope"),
        "isOfficial": bool(batch_payload.get("isOfficial")),
        "officialDate": batch_payload.get("officialDate"),
        "title": batch_payload.get("batchTitle") or "",
        "notes": batch_payload.get("batchNotes") or "",
        "collectedCount": batch_payload.get("collectedCount"),
        "failedCount": batch_payload.get("failedCount"),
    }


def _top_probability(probabilities: dict[str, Any]) -> tuple[str, float]:
    candidates = [(key, _number(probabilities.get(key)) or 0.0) for key in MARKET_KEYS]
    return max(candidates, key=lambda item: item[1])


def _match_label(match: dict[str, Any], row: Any) -> str:
    home = match.get("homeZh") or translate_team_display(match.get("home") or row["home_team"], "主队")
    away = match.get("awayZh") or translate_team_display(match.get("away") or row["away_team"], "客队")
    return f"{home} vs {away}"


def _number(value: Any) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _avg(values: Any) -> float | None:
    numbers = [float(value) for value in values if value is not None]
    return mean(numbers) if numbers else None


def _summary_rows(review: dict[str, Any]) -> list[list[Any]]:
    summary = review.get("summary") or {}
    batch = review.get("batch") or {}
    return [
        ["项目", "值"],
        ["复盘日期", review.get("date")],
        ["复盘来源", "官方批次" if summary.get("isOfficialBatch") else review.get("source")],
        ["批次 ID", batch.get("batchRunId") or "-"],
        ["去重比赛", summary.get("totalMatches")],
        ["已结算", summary.get("settledMatches")],
        ["待结算", summary.get("pendingMatches")],
        ["命中", summary.get("hitCount")],
        ["命中率", summary.get("hitRate")],
        ["平均 Brier", summary.get("avgBrier")],
        ["平均 LogLoss", summary.get("avgLogLoss")],
        ["EV候选", summary.get("evCandidateCount")],
        ["已结算EV候选", summary.get("settledEvCandidateCount")],
        ["EV候选等额盈亏", summary.get("settledEvNetPerUnit")],
        ["高EV实际亏损", summary.get("highEvLossCount")],
        ["高EV异常", summary.get("highEvAnomalyCount")],
        ["比分分布审计", ((review.get("scoreDistributionBacktest") or {}).get("summary") or {}).get("settledMatches")],
        ["平均总进球偏差", ((review.get("scoreDistributionBacktest") or {}).get("summary") or {}).get("avgTotalGoalError")],
        ["正式资金占用", summary.get("formalStake")],
        ["正式信号状态", summary.get("formalSignalState")],
        ["模拟舱现金", summary.get("paperCash")],
        ["模拟舱预留", summary.get("paperReservedStake")],
        ["模拟舱权益", summary.get("paperEquity")],
        ["模拟舱回撤", summary.get("paperDrawdownPct")],
        ["模拟舱风险", summary.get("paperRiskLabel")],
    ]


def _write_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    headers: list[str] | None = None,
) -> None:
    ws = wb.create_sheet(title)
    if headers is None:
        headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="F36B21")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1EA")
    percent_keywords = ("Rate", "Probability", "Value", "Brier", "Loss", "Quality", "Pct")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                header = str(ws.cell(1, cell.column).value or "")
                if isinstance(cell.value, float):
                    if any(keyword in header for keyword in percent_keywords) or header in {"命中率", "最高概率"}:
                        cell.number_format = "0.0%"
                    else:
                        cell.number_format = "0.000"
        for col in range(1, ws.max_column + 1):
            width = 10
            for cell in ws[get_column_letter(col)]:
                width = max(width, min(len(str(cell.value or "")) + 2, 34))
            ws.column_dimensions[get_column_letter(col)].width = width
