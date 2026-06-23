from __future__ import annotations

import math
from io import BytesIO
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from .localization import (
    localize_selection,
    translate_league_display,
    translate_team_display,
)

MODEL_DIVERGENCE_LIMIT = 0.15
SUSPENDED_EV_STATUSES = {"SUSPENDED_MODEL_DIVERGENCE", "MODEL_MARKET_CONFLICT", "SUSPENDED"}


def build_chinese_report(payload: dict[str, Any]) -> str:
    payload = _report_safe_payload(payload)
    match = payload.get("match") or {}
    meta = payload.get("meta") or {}
    probabilities = payload.get("probabilities") or {}
    display = probabilities.get("display") or probabilities.get("final") or {}
    pbase = probabilities.get("pbase") or probabilities.get("model") or {}
    qmkt = probabilities.get("qmkt") or probabilities.get("market") or {}
    governance = payload.get("modelGovernance") or {}
    model_audit = payload.get("modelAudit") or {}
    validation = payload.get("modelValidation") or {}
    expected_goals = payload.get("expectedGoals") or {}
    data_quality = payload.get("dataQuality") or {}
    processing = payload.get("dataProcessing") or {}
    portfolio = payload.get("portfolio") or {}
    recommendations = payload.get("recommendations") or []
    scores = payload.get("topScores") or []
    notes = _localized_notes(payload.get("notes") or [], match)

    home = _localized_team(match, "home", "球队A")
    away = _localized_team(match, "away", "球队B")
    league = _localized_league(meta)
    kickoff = str(meta.get("kickoffBeijing") or meta.get("kickoff") or "-")
    bookmaker_priority, received_bookmaker = _bookmaker_labels(payload)
    generated_at = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S 北京时间")

    lines = [
        f"# 世界杯预测报告：{home} vs {away}",
        "",
        f"- 生成时间：{generated_at}",
        f"- 预测口径：90 分钟赛果",
        f"- 数据来源：{meta.get('dataSource') or '-'}",
        f"- 比赛 ID：{match.get('id') or meta.get('fixtureId') or '-'}",
        f"- 数据快照编号：{payload.get('snapshotId') or meta.get('snapshotId') or '-'}",
        f"- 赛事名称：{league}",
        f"- 开赛时间：{kickoff}",
        f"- 场地：{meta.get('venue') or '-'}",
        f"- 庄家优先级：{bookmaker_priority}",
        f"- 实际盘口庄家：{received_bookmaker}",
        f"- 赔率抓取时间：{meta.get('oddsCapturedAtBeijing') or meta.get('oddsCapturedAt') or '-'}",
        f"- 数据抓取模式：{meta.get('collectionModeZh') or (processing.get('collectionModeZh') if processing else '-')}",
        f"- 技术统计覆盖：{meta.get('deepStatsMatches') if meta.get('deepStatsMatches') is not None else (processing.get('deepStatsMatches') if processing else '-')} 场",
        f"- 近期有效比赛：{home} {meta.get('recentMatchesHome', '-')} 场 / {away} {meta.get('recentMatchesAway', '-')} 场",
        f"- 运行编号：{payload.get('runId') or '-'}",
        f"- 正式 EV 状态：{governance.get('gateLabel') or governance.get('gate_label') or '-'}",
        f"- 模型分歧状态：{model_audit.get('statusLabel') or '-'}",
        f"- pfinal 验收状态：{validation.get('statusLabel') or '尚无独立验收报告'}",
        f"- 合格已结算校准样本：{validation.get('eligibleSamples') if validation else '-'}",
        f"- 基础模型版本：{meta.get('pbaseModelVersion') or validation.get('pbaseVersion') or '-'}",
        f"- 市场数据集版本：{meta.get('marketDatasetVersion') or validation.get('datasetVersion') or '-'}",
        "",
        "## 一、核心概率",
        "",
        "| 结果 | 展示融合概率（非 pfinal） | pbase 基础概率 | qmkt 市场去水概率 |",
        "|---|---:|---:|---:|",
        f"| {home} 胜 | {_pct(display.get('home_win'))} | {_pct(pbase.get('home_win'))} | {_pct_or_dash(qmkt.get('home_win') if qmkt else None)} |",
        f"| 平局 | {_pct(display.get('draw'))} | {_pct(pbase.get('draw'))} | {_pct_or_dash(qmkt.get('draw') if qmkt else None)} |",
        f"| {away} 胜 | {_pct(display.get('away_win'))} | {_pct(pbase.get('away_win'))} | {_pct_or_dash(qmkt.get('away_win') if qmkt else None)} |",
        "",
        "## 二、预期进球与比分",
        "",
        f"- {home} 预期进球：{_num(expected_goals.get('home'))}",
        f"- {away} 预期进球：{_num(expected_goals.get('away'))}",
        "",
        "最可能比分：",
    ]
    exp_summary_rows, exp_component_rows = _exp_audit_rows(payload, home, away)
    if exp_summary_rows:
        lines[-1:-1] = [
            "",
            "### 二附、综合优势 exp 计算",
            "",
            *[f"- {label}：{value}" for label, value in exp_summary_rows],
        ]
        if exp_component_rows:
            lines[-1:-1] = [
                "",
                "| 优势项 | 数值 | 倾向 |",
                "|---|---:|---|",
                *[f"| {label} | {value} | {side} |" for label, value, side in exp_component_rows],
            ]
    if scores:
        for item in scores[:6]:
            lines.append(f"- {item.get('score')}: {_pct(item.get('probability'))}")
    else:
        lines.append("- 暂无")

    lines.extend(
        [
            "",
            "## 三、数据质量与市场完整性",
            "",
            f"- 数据质量评分：{_pct(data_quality.get('score'))}",
            f"- 数据质量等级：{data_quality.get('gradeLabel') or '-'}",
            f"- 研究方向最低质量门槛：{_pct(data_quality.get('minQuality'))}",
            "",
            "| 市场 | 状态 | 盘口 | 说明 |",
            "|---|---|---:|---|",
        ]
    )
    markets = data_quality.get("markets") or []
    if markets:
        for item in markets:
            lines.append(
                "| "
                + f"{item.get('label') or '-'} | "
                + f"{item.get('status_label') or item.get('statusLabel') or item.get('status') or '-'} | "
                + f"{_line(item.get('line'))} | "
                + f"{item.get('details') or '-'} |"
            )
    else:
        lines.append("| - | - | - | - |")

    if processing:
        processing_teams = _localized_processing_teams(processing, match)
        lines.extend(
            [
                "",
                "### 三附、数据处理审计",
                "",
                "| 球队 | 有效场数 | 技术覆盖 | 场均积分 | 场均进球 | 场均失球 | 场均xG | 场均射门 | 场均射正 | 平均控球 | 进攻评分 | 防守评分 |",
                "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
            ]
        )
        for team in processing_teams:
            lines.append(
                f"| {team.get('displayName') or '-'} | {team.get('validCount', '-')} | {team.get('technicalCount', '-')} | "
                f"{_num(team.get('pointsPerGame'))} | {_num(team.get('goalsForAverage'))} | "
                f"{_num(team.get('goalsAgainstAverage'))} | {_num(team.get('xgAverage'))} | "
                f"{_num(team.get('shotsAverage'))} | {_num(team.get('shotsOnTargetAverage'))} | "
                f"{_num(team.get('possessionAverage'))} | {_num(team.get('attackRating'))} | "
                f"{_num(team.get('defenseRating'))} |"
            )
        lines.append("")
        lines.append(f"- 赔率轨迹说明：{(processing.get('oddsTrend') or {}).get('message') or '-'}")

    lines.extend(
        [
            "",
            "## 四、模拟舱",
            "",
            f"- 启动资金：{_money(portfolio.get('bankroll'))}",
            f"- 均注金额：{_money(portfolio.get('unit_stake'))}",
            f"- 本场占用：{_money(portfolio.get('total_stake'))}",
            f"- 模拟期望收益：{_money(portfolio.get('expected_profit'))}",
            f"- 期望资金：{_money(portfolio.get('expected_bankroll'))}",
            "",
                "| 市场 | 方向 | 动作 | 赔率 | 概率口径 | qmkt 市场概率 | 研究试算EV/注 | 纸上EV(p_adj) | EV计算路径 | 说明 |",
            "|---|---|---|---:|---:|---:|---:|---:|---|---|",
        ]
    )
    for item in recommendations:
        lines.append(
            "| "
            + f"{item.get('market') or '-'} | "
            + f"{_localized_selection(item, match)} | "
            + f"{_action_label(item.get('action'))} | "
            + f"{_odd(item.get('odds'))} | "
            + f"{_probability_text(item)} | "
            + f"{_pct_or_dash(item.get('market_probability'))} | "
            + f"{_display_ev(item, 'expected_value_per_unit')} | "
            + f"{_display_ev(item, 'conservative_expected_value_per_unit')} | "
            + f"{_ev_path_text(item)} | "
            + f"{item.get('reason') or '-'} |"
        )

    audit_rows = _audit_rows(recommendations, match)
    if audit_rows:
        lines.extend(
            [
                "",
                "### 模型异常审计附录（原始研究试算，不构成信号）",
                "",
                "| 市场 | 方向 | 原始研究试算EV | 原始纸上试算EV |",
                "|---|---|---:|---:|",
            ]
        )
        for row in audit_rows:
            lines.append(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} |")

    lines.extend(["", "## 五、数据提示", ""])
    if notes:
        lines.extend(f"- {note}" for note in notes)
    else:
        lines.append("- 暂无")

    lines.extend(
        [
            "",
            "## 六、重要边界",
            "",
            "- 本报告只看 90 分钟赛果。",
            "- 展示融合概率仅用于比较，不是已验证的 pfinal。",
            "- 当前纸上模拟按 paper_EV / p_adj 占用纸上资金；pfinal/formal_EV 通过校准回测前，真实执行仍关闭。",
            "- 当基础模型与当前市场基准去水概率发生重大分歧时，主界面与模拟舱暂停 EV 数值；原始试算仅保留在审计附录。",
            "- 当前时间切分校准审计仅覆盖胜平负；大小球和让球仍需独立的比分分布验收。",
            "- 正式 API 模式仅使用庄家优先级中同一全场盘口的赔率计算 EV。",
            "- 本报告用于本地研究和纸上回测，不连接真实投注平台，也不保证收益。",
        ]
    )
    return "\n".join(lines) + "\n"


def build_excel_report(payload: dict[str, Any]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    data = _report_data(payload)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "中文报告"
    sheet.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="F26A21")
    section_fill = PatternFill("solid", fgColor="FFF0E6")
    header_fill = PatternFill("solid", fgColor="F2F4F7")
    white_font = Font(color="FFFFFF", bold=True, size=16)
    section_font = Font(color="9A3412", bold=True, size=13)
    header_font = Font(bold=True, color="667085")
    body_font = Font(name="Arial Unicode MS", size=11)

    row = 1
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = sheet.cell(row=row, column=1, value=f"世界杯预测报告：{data['home']} vs {data['away']}")
    cell.fill = title_fill
    cell.font = white_font
    cell.alignment = Alignment(horizontal="center")
    row += 2

    row = _write_key_values(sheet, row, data["summary"], body_font)
    row = _write_table(sheet, row + 1, "核心概率", ["结果", "展示融合概率（非 pfinal）", "pbase 基础概率", "qmkt 市场去水概率"], data["probability_rows"], section_fill, section_font, header_fill, header_font, body_font)
    row = _write_table(sheet, row + 1, "预期进球与综合优势 exp", ["项目", "数值"], data["expected_rows"], section_fill, section_font, header_fill, header_font, body_font)
    if data["exp_component_rows"]:
        row = _write_table(sheet, row + 1, "综合优势分解", ["优势项", "数值", "倾向"], data["exp_component_rows"], section_fill, section_font, header_fill, header_font, body_font)
    row = _write_table(sheet, row + 1, "最可能比分", ["比分", "概率"], data["score_rows"], section_fill, section_font, header_fill, header_font, body_font)
    row = _write_table(sheet, row + 1, "数据质量与市场完整性", ["市场", "状态", "盘口", "说明"], data["quality_rows"], section_fill, section_font, header_fill, header_font, body_font)
    if data["processing_rows"]:
        row = _write_table(sheet, row + 1, "数据处理审计", ["球队", "有效场数", "技术覆盖", "场均积分", "场均进球", "场均失球", "场均xG", "场均射门", "场均射正", "平均控球", "进攻评分", "防守评分"], data["processing_rows"], section_fill, section_font, header_fill, header_font, body_font)
    if data["recent_match_rows"]:
        row = _write_table(sheet, row + 1, "近期比赛明细", ["球队", "北京时间", "对手", "赛事", "主客", "比分", "xG", "射门", "射正", "控球", "红牌", "点球", "结果"], data["recent_match_rows"], section_fill, section_font, header_fill, header_font, body_font)
    row = _write_table(sheet, row + 1, "模拟舱", ["市场", "方向", "动作", "赔率", "概率口径", "qmkt 市场概率", "研究试算EV/注", "纸上EV(p_adj)", "EV计算路径", "说明"], data["recommendation_rows"], section_fill, section_font, header_fill, header_font, body_font)
    if data["audit_rows"]:
        row = _write_table(sheet, row + 1, "模型异常审计（原始试算，不构成信号）", ["市场", "方向", "原始研究试算EV", "原始纸上试算EV"], data["audit_rows"], section_fill, section_font, header_fill, header_font, body_font)
    _write_table(sheet, row + 1, "数据提示", ["说明"], [[note] for note in data["notes"]], section_fill, section_font, header_fill, header_font, body_font)

    widths = [18, 18, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 42]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for rows in sheet.iter_rows():
        for cell in rows:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_pdf_report(payload: dict[str, Any]) -> bytes:
    from PIL import Image, ImageDraw, ImageFont

    data = _report_data(payload)
    width, height = 1240, 1754
    margin = 70
    line_gap = 12
    orange = (242, 106, 33)
    ink = (23, 32, 42)
    muted = (102, 112, 133)
    line = (217, 222, 232)

    title_font = _load_font(34)
    section_font = _load_font(25)
    header_font = _load_font(20)
    body_font = _load_font(19)
    small_font = _load_font(17)

    pages: list[Image.Image] = []
    page = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(page)
    y = margin

    def ensure(space: int) -> None:
        nonlocal page, draw, y
        if y + space <= height - margin:
            return
        pages.append(page)
        page = Image.new("RGB", (width, height), "white")
        draw = ImageDraw.Draw(page)
        y = margin

    def text_line(text: str, font, fill=ink, gap=line_gap) -> None:
        nonlocal y
        for part in _wrap_pdf_text(draw, str(text), font, width - margin * 2):
            ensure(34)
            draw.text((margin, y), part, font=font, fill=fill)
            y += _font_height(font) + gap

    def section(title: str) -> None:
        nonlocal y
        ensure(58)
        y += 8
        draw.rectangle((margin, y, width - margin, y + 38), fill=(255, 240, 230), outline=(254, 215, 170))
        draw.text((margin + 12, y + 5), title, font=section_font, fill=(154, 52, 18))
        y += 52

    draw.rectangle((0, 0, width, 92), fill=orange)
    draw.text((margin, 24), f"世界杯预测报告：{data['home']} vs {data['away']}", font=title_font, fill="white")
    y = 120

    for key, value in data["summary"]:
        text_line(f"{key}：{value}", body_font)

    section("一、核心概率")
    for row in data["probability_rows"]:
        text_line("  ".join(str(item) for item in row), body_font)

    section("二、预期进球与比分")
    for row in data["expected_rows"]:
        text_line("  ".join(str(item) for item in row), body_font)
    if data["exp_component_rows"]:
        text_line("综合优势分解：" + "；".join(" ".join(str(item) for item in row) for row in data["exp_component_rows"]), small_font, muted)
    text_line("最可能比分：" + "；".join(f"{score} {prob}" for score, prob in data["score_rows"]), small_font, muted)

    section("三、数据质量与市场完整性")
    for row in data["quality_rows"]:
        text_line("  ".join(str(item) for item in row), body_font)
    if data["processing_rows"]:
        section("三附、数据处理审计")
        for row in data["processing_rows"]:
            text_line(" | ".join(str(item) for item in row), small_font)
        text_line(f"赔率轨迹说明：{data['odds_trend_note']}", small_font, muted)

    section("四、模拟舱")
    for row in data["bankroll_rows"]:
        text_line("  ".join(str(item) for item in row), body_font)
    for row in data["recommendation_rows"]:
        text_line(" | ".join(str(item) for item in row), small_font)
    if data["audit_rows"]:
        section("四附、模型异常审计（不构成信号）")
        for row in data["audit_rows"]:
            text_line(" | ".join(str(item) for item in row), small_font)

    section("五、数据提示与边界")
    for note in data["notes"]:
        text_line(f"- {note}", small_font, muted)
    for note in data["boundaries"]:
        text_line(f"- {note}", small_font, muted)

    pages.append(page)
    output = BytesIO()
    pages[0].save(output, "PDF", save_all=True, append_images=pages[1:], resolution=144.0)
    return output.getvalue()


def _report_data(payload: dict[str, Any]) -> dict[str, Any]:
    payload = _report_safe_payload(payload)
    match = payload.get("match") or {}
    meta = payload.get("meta") or {}
    probabilities = payload.get("probabilities") or {}
    display = probabilities.get("display") or probabilities.get("final") or {}
    pbase = probabilities.get("pbase") or probabilities.get("model") or {}
    qmkt = probabilities.get("qmkt") or probabilities.get("market") or {}
    governance = payload.get("modelGovernance") or {}
    model_audit = payload.get("modelAudit") or {}
    validation = payload.get("modelValidation") or {}
    expected_goals = payload.get("expectedGoals") or {}
    data_quality = payload.get("dataQuality") or {}
    portfolio = payload.get("portfolio") or {}
    recommendations = payload.get("recommendations") or []
    scores = payload.get("topScores") or []
    notes = _localized_notes(payload.get("notes") or [], match)
    processing = payload.get("dataProcessing") or {}

    home = _localized_team(match, "home", "球队A")
    away = _localized_team(match, "away", "球队B")
    league = _localized_league(meta)
    kickoff = str(meta.get("kickoffBeijing") or meta.get("kickoff") or "-")
    bookmaker_priority, received_bookmaker = _bookmaker_labels(payload)
    generated_at = datetime.now(ZoneInfo("Asia/Shanghai")).strftime("%Y-%m-%d %H:%M:%S 北京时间")
    quality_markets = data_quality.get("markets") or []
    processing_teams = _localized_processing_teams(processing, match)
    recent_match_rows = []
    for team in processing_teams:
        for item in team.get("matches") or []:
            recent_match_rows.append(
                [
                    team.get("displayName") or "-",
                    item.get("dateBeijing") or "-",
                    item.get("opponentZh") or "-",
                    item.get("leagueZh") or "-",
                    item.get("venueLabel") or "-",
                    f"{item.get('goalsFor', '-')} - {item.get('goalsAgainst', '-')}",
                    _num(item.get("xg")),
                    _num(item.get("shots")),
                    _num(item.get("shotsOnTarget")),
                    _num(item.get("possessionPct")),
                    item.get("redCards") if item.get("redCards") is not None else "-",
                    item.get("penalties") if item.get("penalties") is not None else "-",
                    item.get("resultLabel") or "-",
                ]
            )

    return {
        "home": home,
        "away": away,
        "summary": [
            ("生成时间", generated_at),
            ("预测口径", "90 分钟赛果"),
            ("数据来源", meta.get("dataSource") or "-"),
            ("比赛 ID", match.get("id") or meta.get("fixtureId") or "-"),
            ("数据快照编号", payload.get("snapshotId") or meta.get("snapshotId") or "-"),
            ("赛事名称", league),
            ("开赛时间", kickoff),
            ("场地", meta.get("venue") or "-"),
            ("庄家优先级", bookmaker_priority),
            ("实际盘口庄家", received_bookmaker),
            ("赔率抓取时间", meta.get("oddsCapturedAtBeijing") or meta.get("oddsCapturedAt") or "-"),
            ("数据抓取模式", meta.get("collectionModeZh") or (processing.get("collectionModeZh") if processing else "-")),
            (
                "技术统计覆盖",
                f"{meta.get('deepStatsMatches') if meta.get('deepStatsMatches') is not None else (processing.get('deepStatsMatches') if processing else '-')} 场",
            ),
            (
                "近期有效比赛",
                f"{home} {meta.get('recentMatchesHome', '-')} 场 / {away} {meta.get('recentMatchesAway', '-')} 场",
            ),
            ("运行编号", payload.get("runId") or "-"),
            ("正式 EV 状态", governance.get("gateLabel") or governance.get("gate_label") or "-"),
            ("模型分歧状态", model_audit.get("statusLabel") or "-"),
            ("pfinal 验收状态", validation.get("statusLabel") or "尚无独立验收报告"),
            ("合格已结算校准样本", str(validation.get("eligibleSamples")) if validation else "-"),
            ("基础模型版本", meta.get("pbaseModelVersion") or validation.get("pbaseVersion") or "-"),
            ("市场数据集版本", meta.get("marketDatasetVersion") or validation.get("datasetVersion") or "-"),
        ],
        "probability_rows": [
            [f"{home} 胜", _pct(display.get("home_win")), _pct(pbase.get("home_win")), _pct_or_dash(qmkt.get("home_win") if qmkt else None)],
            ["平局", _pct(display.get("draw")), _pct(pbase.get("draw")), _pct_or_dash(qmkt.get("draw") if qmkt else None)],
            [f"{away} 胜", _pct(display.get("away_win")), _pct(pbase.get("away_win")), _pct_or_dash(qmkt.get("away_win") if qmkt else None)],
        ],
        "expected_rows": _expected_rows(payload, home, away),
        "exp_component_rows": _exp_audit_rows(payload, home, away)[1],
        "score_rows": [[item.get("score"), _pct(item.get("probability"))] for item in scores[:8]] or [["暂无", "-"]],
        "quality_rows": [
            [
                item.get("label") or "-",
                item.get("status_label") or item.get("statusLabel") or item.get("status") or "-",
                _line(item.get("line")),
                item.get("details") or "-",
            ]
            for item in quality_markets
        ]
        or [["-", "-", "-", "-"]],
        "processing_rows": [
            [
                team.get("displayName") or "-",
                team.get("validCount", "-"),
                team.get("technicalCount", "-"),
                _num(team.get("pointsPerGame")),
                _num(team.get("goalsForAverage")),
                _num(team.get("goalsAgainstAverage")),
                _num(team.get("xgAverage")),
                _num(team.get("shotsAverage")),
                _num(team.get("shotsOnTargetAverage")),
                _num(team.get("possessionAverage")),
                _num(team.get("attackRating")),
                _num(team.get("defenseRating")),
            ]
            for team in processing_teams
        ],
        "recent_match_rows": recent_match_rows,
        "odds_trend_note": (processing.get("oddsTrend") or {}).get("message") or "-",
        "bankroll_rows": [
            ["启动资金", _money(portfolio.get("bankroll"))],
            ["均注金额", _money(portfolio.get("unit_stake"))],
            ["本场占用", _money(portfolio.get("total_stake"))],
            ["模拟期望收益", _money(portfolio.get("expected_profit"))],
            ["期望资金", _money(portfolio.get("expected_bankroll"))],
            ["数据质量评分", _pct(data_quality.get("score"))],
            ["数据质量等级", data_quality.get("gradeLabel") or "-"],
        ],
        "recommendation_rows": [
            [
                item.get("market") or "-",
                _localized_selection(item, match),
                _action_label(item.get("action")),
                _odd(item.get("odds")),
                _probability_text(item),
                _pct_or_dash(item.get("market_probability")),
                _display_ev(item, "expected_value_per_unit"),
                _display_ev(item, "conservative_expected_value_per_unit"),
                _ev_path_text(item),
                item.get("reason") or "-",
            ]
            for item in recommendations
        ]
        or [["-", "-", "-", "-", "-", "-", "-", "-", "-", "-"]],
        "audit_rows": _audit_rows(recommendations, match),
        "notes": list(notes) if notes else ["暂无"],
        "boundaries": [
            "本报告只看 90 分钟赛果。",
            "展示融合概率仅用于比较，不是已验证的 pfinal。",
            "当前纸上模拟按 paper_EV / p_adj 占用纸上资金；pfinal/formal_EV 通过校准回测前，真实执行仍关闭。",
            "基础模型与当前市场基准去水概率重大分歧时，模拟舱暂停 EV 数值；原始试算仅保留在审计附录。",
            "当前时间切分校准审计仅覆盖胜平负；大小球和让球仍需独立的比分分布验收。",
            "正式 API 模式仅使用庄家优先级中同一全场盘口的赔率计算 EV。",
            "本报告用于本地研究和纸上回测，不连接真实投注平台，也不保证收益。",
        ],
    }


def _expected_rows(payload: dict[str, Any], home: str, away: str) -> list[list[str]]:
    expected_goals = payload.get("expectedGoals") or {}
    rows = [
        [f"{home} 最终预期进球 λ", _num(expected_goals.get("home"))],
        [f"{away} 最终预期进球 λ", _num(expected_goals.get("away"))],
    ]
    exp_summary_rows, _ = _exp_audit_rows(payload, home, away)
    rows.extend([[label, value] for label, value in exp_summary_rows])
    return rows


def _exp_audit_rows(payload: dict[str, Any], home: str, away: str) -> tuple[list[tuple[str, str]], list[list[str]]]:
    expected_goals = payload.get("expectedGoals") or {}
    feature_edges = payload.get("featureEdges") or {}
    if not expected_goals and not feature_edges:
        return [], []

    component_keys = [
        "elo_edge",
        "fifa_rank_edge",
        "host_edge",
        "rest_edge",
        "travel_edge",
        "group_context_edge",
        "rotation_edge",
        "h2h_edge",
        "country_relation_edge",
        "commercial_incentive_edge",
    ]
    component_labels = {
        "elo_edge": "Elo 强度差",
        "fifa_rank_edge": "FIFA 排名差",
        "host_edge": "主场/中立场",
        "rest_edge": "休息天数",
        "travel_edge": "旅行距离",
        "group_context_edge": "积分/战意",
        "rotation_edge": "轮换风险",
        "h2h_edge": "历史交锋",
        "country_relation_edge": "国家关系",
        "commercial_incentive_edge": "商业动机",
    }
    components: list[tuple[str, float, str]] = []
    for key in component_keys:
        value = _finite_float(feature_edges.get(key), 0.0)
        side = home if value > 0 else away if value < 0 else "中性"
        components.append((component_labels.get(key, key), value, side))

    fallback_log_edge = sum(value for _, value, _ in components)
    log_edge = _finite_float(expected_goals.get("logEdge"), fallback_log_edge)
    home_multiplier = _finite_float(expected_goals.get("homeExpMultiplier"), math.exp(log_edge))
    away_multiplier = _finite_float(expected_goals.get("awayExpMultiplier"), math.exp(-log_edge))
    draw_boost = _finite_float(feature_edges.get("rivalry_draw_boost"), 0.0)

    summary_rows: list[tuple[str, str]] = [
        ("基础 λ", f"{home} {_num(expected_goals.get('baseHome'))} / {away} {_num(expected_goals.get('baseAway'))}"),
        ("综合优势 log_edge", f"{_signed_num(log_edge, 3)}（正数偏向 {home}，负数偏向 {away}）"),
        ("exp 倍率", f"{home} ×{_num(home_multiplier, 3)} / {away} ×{_num(away_multiplier, 3)}"),
        ("exp 修正后 λ", f"{home} {_num(expected_goals.get('rawHome'))} / {away} {_num(expected_goals.get('rawAway'))}"),
        (
            "风险收缩后最终 λ",
            f"{home} {_num(expected_goals.get('home'))} / {away} {_num(expected_goals.get('away'))}，shrink_factor {_num(expected_goals.get('lambdaShrinkFactor'))}",
        ),
        ("公式", f"{home} λ = 基础 λ × exp(log_edge)；{away} λ = 基础 λ × exp(-log_edge)"),
        ("平局增强", f"draw_boost {_signed_num(draw_boost, 3)}，不进入 exp，单独修正平局概率"),
    ]
    component_rows = [
        [label, _signed_num(value, 3), side]
        for label, value, side in components
        if abs(value) > 1e-12
    ]
    return summary_rows, component_rows


def _write_key_values(sheet, row, values, body_font):
    from openpyxl.styles import Font

    for key, value in values:
        sheet.cell(row=row, column=1, value=key).font = Font(bold=True, color="667085")
        sheet.cell(row=row, column=2, value=value).font = body_font
        row += 1
    return row


def _localized_team(match: dict[str, Any], side: str, fallback: str) -> str:
    original = str(match.get(side) or "").strip()
    displayed = str(match.get(f"{side}Zh") or "").strip()
    if displayed and not _looks_untranslated(displayed) and "名称待核定" not in displayed:
        return displayed
    translated = translate_team_display(original or displayed, fallback)
    return translated or displayed or original or fallback


def _localized_league(meta: dict[str, Any]) -> str:
    original = str(meta.get("leagueName") or "").strip()
    displayed = str(meta.get("leagueNameZh") or "").strip()
    country = str(meta.get("leagueCountry") or "").strip()
    if displayed and not _looks_untranslated(displayed) and "名称待核定" not in displayed:
        return displayed
    translated = translate_league_display(original or displayed, country)
    return translated or displayed or original or "-"


def _localized_processing_teams(processing: dict[str, Any], match: dict[str, Any]) -> list[dict[str, Any]]:
    teams: list[dict[str, Any]] = []
    for side, fallback in (("home", "主队"), ("away", "客队")):
        raw_team = processing.get(side) or {}
        if not raw_team:
            continue
        team = deepcopy(raw_team)
        display_name = str(team.get("displayName") or "").strip()
        original_name = str(team.get("rawName") or match.get(side) or "").strip()
        if not display_name or _looks_untranslated(display_name) or "名称待核定" in display_name:
            team["displayName"] = translate_team_display(original_name or display_name, fallback)
        localized_matches: list[dict[str, Any]] = []
        for raw_item in team.get("matches") or []:
            item = deepcopy(raw_item)
            opponent_display = str(item.get("opponentZh") or "").strip()
            if not opponent_display or "名称待核定" in opponent_display:
                item["opponentZh"] = translate_team_display(item.get("opponent"), "对手")
            league_display = str(item.get("leagueZh") or "").strip()
            if not league_display or "名称待核定" in league_display:
                item["leagueZh"] = translate_league_display(item.get("league"))
            localized_matches.append(item)
        team["matches"] = localized_matches
        teams.append(team)
    return teams


def _localized_notes(notes: list[Any], match: dict[str, Any]) -> list[str]:
    replacements = {
        "主队（名称待核定）": _localized_team(match, "home", "主队"),
        "客队（名称待核定）": _localized_team(match, "away", "客队"),
    }
    localized: list[str] = []
    for note in notes:
        text = str(note)
        for source, target in replacements.items():
            text = text.replace(source, target)
        localized.append(text)
    return localized


def _looks_untranslated(value: str) -> bool:
    return any("A" <= char <= "Z" or "a" <= char <= "z" for char in value)


def _write_table(sheet, row, title, headers, rows, section_fill, section_font, header_fill, header_font, body_font):
    sheet.cell(row=row, column=1, value=title).fill = section_fill
    sheet.cell(row=row, column=1).font = section_font
    row += 1
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font
    row += 1
    for item in rows:
        for column, value in enumerate(item, start=1):
            sheet.cell(row=row, column=column, value=value).font = body_font
        row += 1
    return row


def _load_font(size: int):
    from PIL import ImageFont

    candidates = [
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
    ]
    for path in candidates:
        if not Path(path).exists():
            continue
        try:
            return ImageFont.truetype(path, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def _font_height(font) -> int:
    bbox = font.getbbox("中文报告")
    return max(1, bbox[3] - bbox[1])


def _wrap_pdf_text(draw, text: str, font, max_width: int) -> list[str]:
    if not text:
        return [""]
    lines: list[str] = []
    current = ""
    for char in text:
        trial = current + char
        if current and draw.textlength(trial, font=font) > max_width:
            lines.append(current)
            current = char
        else:
            current = trial
    if current:
        lines.append(current)
    return lines


def _pct(value: Any) -> str:
    try:
        return f"{float(value) * 100:.1f}%"
    except (TypeError, ValueError):
        return "-"


def _probability_text(item: dict[str, Any]) -> str:
    label = str(item.get("model_probability_label") or "").strip()
    value = _pct(item.get("model_probability"))
    if not label:
        label = "模型胜率" if item.get("market") == "胜平负" else "正收益概率"
    return f"{label} {value}"


def _pct_or_dash(value: Any) -> str:
    return _pct(value)


def _display_ev(item: dict[str, Any], key: str) -> str:
    if item.get("ev_status") in SUSPENDED_EV_STATUSES:
        return "暂停"
    return _pct_or_dash(item.get(key))


def _ev_path_text(item: dict[str, Any]) -> str:
    if item.get("ev_status") in SUSPENDED_EV_STATUSES:
        return "模型分歧超限，EV 已暂停；原始试算仅供审计。"
    calc = item.get("ev_calculation") or {}
    if not isinstance(calc, dict) or not calc.get("formula"):
        return "-"

    calc_type = str(calc.get("type") or "")
    odds = _odd(calc.get("odds"))
    ev = _pct_or_dash(calc.get("expectedValue"))
    paper_ev = _pct_or_dash(calc.get("paperExpectedValue") if calc.get("paperExpectedValue") is not None else calc.get("conservativeExpectedValue"))
    p_adj = _pct_or_dash(calc.get("adjustedProbability"))
    shrink_k = _num(calc.get("shrinkK"))
    if calc_type == "1X2":
        return (
            f"EV = pbase {_pct(calc.get('modelProbability'))} × 赔率 {odds} - 1 = {ev}；"
            f"展开为盈利概率 {_pct(calc.get('modelProbability'))} × 净赔付 "
            f"{_num((float(calc.get('odds') or 0.0) - 1.0) if calc.get('odds') else None)} "
            f"- 亏损概率 {_pct(calc.get('lossStakeFraction'))}；"
            f"p_adj {p_adj}，shrink_k {shrink_k}；纸上EV {paper_ev}。"
        )

    settlement = calc.get("settlement") or {}
    calibration = calc.get("scoreDistributionCalibration") or calc.get("totalGoalsCalibration") or {}
    calibration_text = ""
    if isinstance(calibration, dict) and calibration.get("applied"):
        calibration_text = (
            f"分布校准：{calibration.get('marketLabel') or '比分分布'}·{calibration.get('sideLabel') or '方向'}"
            f"因子 {_num(calibration.get('positiveFactor'))}，"
            f"样本 {calibration.get('sampleCount') or 0}，"
            f"原始EV {_pct_or_dash(calc.get('rawExpectedValue'))} -> 校准EV {ev}；"
        )
    return (
        f"{calc.get('formula') or 'EV'}：盈利注权重 {_num(calc.get('winStakeFraction'))} × "
        f"(赔率 {odds} - 1) - 亏损注权重 {_num(calc.get('lossStakeFraction'))} = {ev}；"
        f"{calibration_text}"
        f"全赢 {_pct(settlement.get('fullWinProbability'))}，半赢 {_pct(settlement.get('halfWinProbability'))}，"
        f"走水 {_pct(settlement.get('pushProbability'))}，半输 {_pct(settlement.get('halfLossProbability'))}，"
        f"全输 {_pct(settlement.get('fullLossProbability'))}；"
        f"盈亏平衡赔率 {_odd(calc.get('breakEvenOdds'))}；p_adj {p_adj}，shrink_k {shrink_k}；纸上EV {paper_ev}。"
    )


def _audit_rows(recommendations: list[dict[str, Any]], match: dict[str, Any]) -> list[list[str]]:
    return [
        [
            str(item.get("market") or "-"),
            _localized_selection(item, match),
            _pct_or_dash(item.get("audit_expected_value_per_unit")),
            _pct_or_dash(item.get("audit_paper_expected_value_per_unit") if item.get("audit_paper_expected_value_per_unit") is not None else item.get("audit_conservative_expected_value_per_unit")),
        ]
        for item in recommendations
        if item.get("ev_status") in SUSPENDED_EV_STATUSES
        and item.get("audit_expected_value_per_unit") is not None
    ]


def _localized_selection(item: dict[str, Any], match: dict[str, Any]) -> str:
    selection = str(item.get("selection") or "-")
    return localize_selection(selection, str(match.get("home") or ""), str(match.get("away") or ""))


def _report_safe_payload(payload: dict[str, Any]) -> dict[str, Any]:
    if (payload.get("modelAudit") or {}).get("evSuspended"):
        return payload
    benchmark = _bookmaker_for_safe_audit(payload)
    probabilities = payload.get("probabilities") or {}
    pbase = probabilities.get("pbase") or probabilities.get("model") or {}
    qmkt = probabilities.get("qmkt") or probabilities.get("market") or {}
    keys = ("home_win", "draw", "away_win")
    if not all(key in pbase and key in qmkt for key in keys):
        return payload
    labels = {"home_win": "主胜", "draw": "平局", "away_win": "客胜"}
    gaps = {key: float(pbase[key]) - float(qmkt[key]) for key in keys}
    trigger_key, signed_gap = max(gaps.items(), key=lambda item: abs(item[1]))
    max_gap = abs(signed_gap)
    if max_gap <= MODEL_DIVERGENCE_LIMIT:
        return payload

    safe = deepcopy(payload)
    reason = (
        f"模型分歧异常：{labels[trigger_key]} 的 pbase 与 {benchmark} 去水概率相差 "
        f"{max_gap * 100:.1f} 个百分点，超过 {MODEL_DIVERGENCE_LIMIT * 100:.1f} 个百分点；"
        "本场所有市场 EV 暂停计算，仅供模型复核。"
    )
    safe["modelAudit"] = {
        "status": "ANOMALY",
        "statusLabel": "模型分歧异常",
        "evSuspended": True,
        "triggerSelection": labels[trigger_key],
        "maxProbabilityGap": max_gap,
        "threshold": MODEL_DIVERGENCE_LIMIT,
        "reason": reason,
        "gaps": gaps,
    }
    for item in safe.get("recommendations") or []:
        if item.get("expected_value_per_unit") is None:
            continue
        item["audit_expected_value_per_unit"] = item.get("expected_value_per_unit")
        item["audit_conservative_expected_value_per_unit"] = item.get("conservative_expected_value_per_unit")
        item["audit_paper_expected_value_per_unit"] = (
            item.get("paper_expected_value_per_unit") or item.get("conservative_expected_value_per_unit")
        )
        item["ev_pbase_research"] = item.get("ev_pbase_research") or item.get("expected_value_per_unit")
        item["conservative_ev_pbase_research"] = (
            item.get("conservative_ev_pbase_research")
            or item.get("conservative_expected_value_per_unit")
        )
        item["ev_pfinal_exec"] = None
        item["signal_status"] = "SUSPENDED"
        item["expected_value_per_unit"] = None
        item["conservative_expected_value_per_unit"] = None
        item["paper_expected_value_per_unit"] = None
        item["ev_status"] = "MODEL_MARKET_CONFLICT"
        item["decision_status"] = "MODEL_MARKET_CONFLICT"
        item["action"] = "WATCH"
        item["stake"] = 0.0
        item["reason"] = reason
    portfolio = safe.get("portfolio") or {}
    bankroll = portfolio.get("bankroll")
    if bankroll is not None:
        portfolio.update(
            {
                "active_bets": 0,
                "total_stake": 0.0,
                "bankroll_after_stakes": bankroll,
                "expected_profit": 0.0,
                "expected_bankroll": bankroll,
            }
        )
    existing_notes = list(safe.get("notes") or [])
    safe["notes"] = [reason, *[note for note in existing_notes if note != reason]]
    return safe


def _finite_float(value: Any, default: float | None = None) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if not math.isfinite(number):
        return default
    return number


def _num(value: Any, digits: int = 2) -> str:
    number = _finite_float(value)
    if number is None:
        return "-"
    return f"{number:.{digits}f}"


def _signed_num(value: Any, digits: int = 2) -> str:
    number = _finite_float(value)
    if number is None:
        return "-"
    prefix = "+" if number > 0 else ""
    return f"{prefix}{number:.{digits}f}"


def _money(value: Any) -> str:
    number = _finite_float(value)
    if number is None:
        return "-"
    return f"{number:.2f} 元"


def _odd(value: Any) -> str:
    number = _finite_float(value)
    if number is None:
        return "-"
    return f"{number:.2f}"


def _line(value: Any) -> str:
    if value in {None, ""}:
        return "-"
    try:
        return f"{float(value):g}"
    except (TypeError, ValueError):
        return str(value)


def _bookmaker_labels(payload: dict[str, Any]) -> tuple[str, str]:
    meta = payload.get("meta") or {}
    market = payload.get("market") or {}
    priority = meta.get("bookmakerPriority") or market.get("bookmakerPriority")
    if isinstance(priority, list):
        required = " > ".join(str(item) for item in priority if str(item).strip()) or "-"
    else:
        required = str(priority or meta.get("requiredBookmaker") or market.get("requiredBookmaker") or "-")
    selected = market.get("selectedBookmakers") or meta.get("selectedBookmakers")
    if isinstance(selected, dict) and selected:
        labels = {"1X2": "胜平负", "OU": "大小球", "AH": "让球"}
        received = "，".join(f"{labels.get(key, key)}={value}" for key, value in selected.items())
    elif "selectedBookmaker" in market:
        received = market.get("selectedBookmaker") or "未取得"
    else:
        received = meta.get("bookmaker") or "未取得"
    return str(required), str(received)


def _bookmaker_for_safe_audit(payload: dict[str, Any]) -> str:
    meta = payload.get("meta") or {}
    market = payload.get("market") or {}
    selected = market.get("selectedBookmakers") or meta.get("selectedBookmakers")
    if isinstance(selected, dict):
        return str(selected.get("1X2") or market.get("selectedBookmaker") or meta.get("bookmaker") or "当前市场基准")
    return str(market.get("selectedBookmaker") or meta.get("bookmaker") or "当前市场基准")


def _action_label(value: Any) -> str:
    return {
        "BUY": "模型候选（未执行）",
        "PAPER_BUY": "纸上模拟",
        "WATCH": "观望",
        "NO_MARKET": "市场缺失",
        "MODEL_CANDIDATE": "模型候选（未执行）",
        "RESEARCH_WATCH": "研究观察",
        "SUSPENDED": "暂停复核",
    }.get(str(value or ""), str(value or "-"))
